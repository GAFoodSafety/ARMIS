
# coding: utf-8

# In[ ]:

#This is the latest version of ARMIS - Automated Recall Monitoring Information System - as of 5/15/2017

#Packages
import feedparser
from openpyxl import Workbook
from lxml import html
import requests
import json
import nltk
import Tkinter as tk
import csv
from bs4 import BeautifulSoup
from dateutil.parser import *


wb = Workbook()


#Setup Excel Sheet    
ws = wb.active
ws["A1"] = "FDA/USDA"
ws["B1"] = "Class"
ws["C1"] = "Distribution (GA?)"
ws["D1"] = "Date"
ws["E1"] = "Firm"
ws["F1"] = "Contact"
ws["G1"] = "Item"
ws["H1"] = "Reason"
ws["I1"] = "Type"
ws["J1"] = "URL"
ws["K1"] = "Summary Information"

#Retrieving data from RSS and Beginning Variables

URLS = ["http://www.fda.gov/AboutFDA/ContactFDA/StayInformed/RSSFeeds/FoodSafety/rss.xml", 
            "https://www.fsis.usda.gov/wps/wcm/connect/fsis-content/rss/recalls"
]

#listmaker compiles the URLs from both FDA and USDA RSS Feed Sources. The CDC combined Food Safety Recalls is not reliable.
def listmaker(urls):
    global dictA
    dictA = {}
    global keys
    keys = []
    x = feedparser.parse(urls[0])
    y = feedparser.parse(urls[1])
    i = 0
    while i <= len(x):
        dictA.update({str(parse(x.entries[i].published)): ["FDA", x.entries[i].link, x.entries[i].title, x.entries[i].published, x.entries[i].summary, x.entries[i].summary_detail]})
        keys.append(str(parse(x.entries[i].published)))
        i += 1
    j = 0
    while j <= len(y):
        dictA.update({str(parse(y.entries[j].published)): ["USDA", y.entries[j].link, y.entries[j].title, y.entries[j].published, y.entries[j].summary, y.entries[j].summary_detail]})
        keys.append(str(parse(y.entries[j].published)))
        j += 1
    keys = sorted(keys)
    return dictA


#Forming FDA API for information verification. The FDA API is not validated, so this data only is used to cross-check results from RSS feedparsing and HTML scraping.
'''
def apibuilder():
    api = raw_input('Distribution pattern: ')
    api2, api3 = raw_input('Beginning and Ending Recall Dates -> format YYYYMMDD YYYYMMDD: ').split()
    global api4
    api4 = raw_input('Limit: ')
    global url
    url = "https://api.fda.gov/food/enforcement.json?api_key=2WWvb8KRbchVd55OAgPqKRATA8oa0u4XiS0BWOHw&search=(results.distribution_pattern=" + api + ")+AND+(results.recall_initiation_date=[" 
    url = url+ str(api2) 
    url = url + "+TO+" 
    url = url + str(api3)
    url = url + '].exact)&limit=' + str(api4) + "&skip=0"
    print url
apibuilder() 



for i in range(int(api4)):
    classification = (jsonresponse.get("results")[i].get("classification"))
    date = (jsonresponse.get("results")[i].get("recall_initiation_date"))
    firm = (jsonresponse.get("results")[i].get("recalling_firm"))
    origin = (jsonresponse.get("results")[i].get("state")) #(ga facility?)
    product = (jsonresponse.get("results")[i].get("product_description")) #(category)
    reason = (jsonresponse.get("results")[i].get("reason_for_recall"))
    dist = (jsonresponse.get("results")[i].get("distribution_pattern"))
    print "classification" 
    print classification
    print "date" 
    print date
    print "firm" 
    print firm
    print "origin" 
    print origin
    print "product" 
    print product
    print "reason" 
    print reason
    print "distribution" 
    print dist

'''


global app

def quit():
    root.destroy()

listmaker(URLS)

def Soupy(x):
    page = requests.get(dictA[keys[x]][1])
    global soup
    soup = BeautifulSoup(page.content, "lxml")
    global findz
    findz = soup.find_all("p")

            
    
    
#Calling States    
listC = []    
with open('/Users/thomasburker/Documents/Recalls/Corpus/states.csv', 'rU') as csvfile:
    statereader = csv.reader(csvfile, delimiter=' ', quotechar='|')
    for row in statereader:
        #print ', '.join(row)
        listC.append(' '.join(row))

#Calling Types        
listD = []
with open ('/Users/thomasburker/Documents/Recalls/Corpus/Types.csv', 'rU') as csvfile:
    typereader = csv.reader(csvfile, delimiter = ' ', quotechar = '|')
    for row in typereader:
        listD.append(' '.join(row))

#Calling Food        
listX = []
with open ('/Users/thomasburker/Documents/Recalls/Corpus/food.csv', 'rU') as csvfile:
    typereader = csv.reader(csvfile, delimiter = ' ', quotechar = '|')
    for row in typereader:
        listX.append(' '.join(row))
        
def soup_find(x):
    Soupy(x)
    global listE
    listE = []
    global listF
    listF = []
    global listY
    listY = []
    for i in listC:    
        if i in str(findz): 
            listE.append(''.join(i))
    for j in listD:
        if j in str(findz):
            listF.append(''.join(j))
    for zz in listX:
        if zz in str(findz):
            listY.append(''.join(zz))
    listE = sorted(list(set(listE)))
    listF = sorted(list(set(listF)))
    print listE
    print listF
    
    #ws["C" + ] = listE
    #ws["I" + ] = listF

    
            
def trees(result):
    global firm
    firm = []
    for subtree in result.subtrees():
        if subtree.label() == "NP":
            print subtree
            firm.append(subtree)
    
class App:


    def __init__(self, master):


        frame = tk.Frame(master)
        frame.pack()

        self.quitbutton = tk.Button(
            frame, text="QUIT", fg="blue", command=quit
            )
        self.quitbutton.pack()

        self.hi_there = tk.Button(
            frame, text="Initiate ARMIS", command=self.say_hi
            )
        
        self.hi_there.pack()
        S = tk.Scrollbar(root)
        
        T = tk.Text(
            root, height = 20, width = 100
            )
        T.pack()
        S.pack(side=tk.RIGHT, fill=tk.Y)
        T.pack(side=tk.LEFT, fill=tk.Y)
        S.config(command=T.yview)
        T.config(yscrollcommand=S.set)
        
        NA = 0
        MA = 0
        
        
        ij = 0
        #Needs to search the XML/HTML scraped
        while ij < len(keys):
            if ("Georgia" or "GA" or "Ga" or "nationwide" or "50") in dictA[keys[ij]][4] == -1:
                T.insert(tk.END, "Not Georgia\n")
                T.insert(tk.END, dictA[keys[ij]][4])
                T.insert(tk.END, "\n")
            else:
                T.insert(tk.END, "Georgia Outbreak\n")
                T.insert(tk.END, dictA[keys[ij]][4])
                T.insert(tk.END, "\n")
            ij += 1

        #T.insert(END, data )

        
    #def quit_pressed(self):
    #    quit()

    def say_hi(self):
        L = 2
        N = 0
        M = len(keys)

        while N < M:
            V = str(L)
            soup_find(N)
            
            F = dictA[keys[N]]
        #   Labels for Workbook 
            
            #Scraping HTML
            page = requests.get(F[1])
            tree = html.fromstring(page.content)
            data = tree.xpath('//p/text()')
            #Determining Information
            #Agency Determination
            #Use this to dictate action
            ws["A"+ V] = F[0]
            
            #Determine Recall Class
            if F[0] == 'USDA':
                ws["B" + V] = str(soup.find_all("span", attrs={'class':"recall-class-container"}))
                
            #Distribution    
            ws["C"+ V] = str(listE)
            
            #Publish Date
            ws["D"+ V] = keys[N]
            
            #Firm Information
            #Use NLTK
            data4t = nltk.pos_tag(nltk.word_tokenize(F[4]))
            
            grammar = r"""
                NP: {<NNP>+<NN|NNS>+}
                    {<NNP>+<NNP>+}
                """

            
            #cp = nltk.RegexpParser(grammar)
            #result = cp.parse(data4t)
            #trees(result)
            listFirm =[]
            for a,b in data4t:
                listFirm.append(a)
            ws["E"+V] = str(listFirm)
            
            #Contact Information
                #ws["F"+ L] = data[3]
 
            #Item Information
                #ws["G" + V] = ...
                
            #Reason
            if ("allergy" or "allergen" or "allergic" or "undeclared" or "Allergy") in str(findz):
                ws["H"+ V] = "allergy"
                
            elif ("pathogen" or "organism" or "salmonella" or "Salmonella") in str(findz):
                ws["H" + V] = "pathogen"
                
            else:
                ws["H" + V] = "unknown contaminant"
            
            
            #Type
            ws["I"+ V] = str(listF)
            
            #URL
            ws["J"+ V] = F[1]
            
            #Summary Data
            ws["K"+ V] = F[4]

            L += 1
            N += 1   
            wb.save("attempt 3.xlsx")
            

#print getfromdict(F,T)
    
root = tk.Tk()

#w = Canvas(root)
#w.grid()
      
#bindfff


#root.destroy()
app = App(root)

root.mainloop()
#root.destroy()


# In[ ]:




# In[ ]:



